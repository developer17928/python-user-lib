import json
import pandas as pd
from openpyxl import load_workbook


def main():
    # df = pd.read_excel("K-6_Weekly_Book_Schedule.xlsx")
    # df.to_json('output.json', orient='records', indent=4)

    """
    Why not plain pd.read_excel(): the Month and Monthly Theme columns are merged
    across each month's 4 week-rows. openpyxl (and therefore pandas) only stores
    the value in the top-left cell of a merged range -- every other cell in that
    range reads back as None. A naive df.to_json() would leave those cells null
    for weeks 2-4 of every month.

    This module fixes that at the openpyxl level (using the actual MergedCellRange
    objects, not a blind fill) before handing off to pandas, so it's correct even
    if a real blank cell ever ends up next to a merged one.
    """
    SRC = "K-6_Weekly_Book_Schedule.xlsx"

    df = read_merged_sheet_as_df(SRC, sheet_name="Weekly Book Schedule")

    # Sanity check: no more nulls in Month/Theme after unmerging
    assert df["Month"].isna().sum() == 0, "Month column still has unresolved merges"
    assert df["Monthly Theme"].isna().sum() == 0, (
        "Theme column still has unresolved merges"
    )

    nested = schedule_to_nested_json(df)
    flat = schedule_to_flat_records(df)

    with open("K-6_Weekly_Book_Schedule.json", "w") as f:
        json.dump(nested, f, indent=4)

    with open("K-6_Weekly_Book_Schedule_flat.json", "w") as f:
        json.dump(flat, f, indent=4)

    print(f"Rows parsed: {len(df)}")
    print(
        f"Months: {len(nested)}, weeks per month: {[len(m['weeks']) for m in nested]}"
    )
    print("Sample nested entry:")
    print(json.dumps(nested[0], indent=2)[:600])


def read_merged_sheet_as_df(path: str, sheet_name: str = None) -> pd.DataFrame:
    """
    Read an Excel sheet into a DataFrame, resolving merged cells by filling
    every cell in a merged range with that range's top-left value.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Map every merged range to its anchor (top-left) value, then write that
    # value into every cell the merge covers. Do this on a plain grid so we
    # don't fight openpyxl's read-only MergedCell objects.
    grid = [[cell.value for cell in row] for row in ws.iter_rows()]

    for merged_range in ws.merged_cells.ranges:
        anchor_value = grid[merged_range.min_row - 1][merged_range.min_col - 1]
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                grid[row_idx - 1][col_idx - 1] = anchor_value

    header, *data_rows = grid
    return pd.DataFrame(data_rows, columns=header)


def schedule_to_nested_json(df: pd.DataFrame) -> list:
    """
    Group the flat (now fully-populated) DataFrame into a nested structure:
    month -> theme -> weeks -> per-grade books. This mirrors the workbook's
    actual shape better than a flat record list would.
    """
    grade_cols = [
        "Kindergarten",
        "Grade 1",
        "Grade 2",
        "Grade 3",
        "Grade 4",
        "Grade 5",
        "Grade 6",
    ]

    months = []
    for month_name, month_df in df.groupby("Month", sort=False):
        theme = month_df["Monthly Theme"].iloc[0]
        weeks = []
        for _, row in month_df.iterrows():
            weeks.append(
                {
                    "week": row["Week"],
                    "books": {grade: row[grade] for grade in grade_cols},
                }
            )
        months.append({"month": month_name, "theme": theme, "weeks": weeks})

    # groupby(sort=False) preserves first-seen order, but re-sort explicitly
    # by original row order just to be safe against any pandas version quirks
    order = list(dict.fromkeys(df["Month"]))
    months.sort(key=lambda m: order.index(m["month"]))
    return months


def schedule_to_flat_records(df: pd.DataFrame) -> list:
    """Flat version: one JSON object per week/grade cell isn't useful here since
    each row already has all 7 grades -- this just returns one object per
    week-row, which is the direct equivalent of df.to_dict('records')."""
    return json.loads(df.to_json(orient="records"))


if __name__ == "__main__":
    main()
